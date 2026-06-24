import xlrd
import openpyxl
import os
import sys

sys.stdout = open(r"C:\Users\kevin\Desktop\Taospace\excel_output.txt", "w", encoding="utf-8")

BASE = r"C:\Users\kevin\Desktop\Taospace\Taofiles\班務組"

files = {
    "112_南屏": (BASE + r"\112排課\112課表(新民至善-南屏區)0810.xlsx", "xlsx"),
    "113_正慈新民": (BASE + r"\113排課\113正慈新民班.xlsx", "xlsx"),
    "113_正慈至善": (BASE + r"\113排課\113正慈至善.xlsx", "xlsx"),
    "113_正儀新民": (BASE + r"\113排課\113正儀新民班.xlsx", "xlsx"),
    "114_南屏": (BASE + r"\114排課\114課表(新民至善-南屏區)0830.xls", "xls"),
}

def read_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n--- 工作表：{name} ---")
        for row in ws.iter_rows(values_only=True):
            vals = [str(v).strip() for v in row if v is not None and str(v).strip() not in ('', 'None')]
            if vals:
                print(" | ".join(vals))

def read_xls(path):
    for enc in ['big5', 'cp950', 'utf-8', 'gbk']:
        try:
            wb = xlrd.open_workbook(path, encoding_override=enc)
            for sheet in wb.sheets():
                print(f"\n--- 工作表：{sheet.name} (encoding={enc}) ---")
                for row in range(sheet.nrows):
                    vals = [str(sheet.cell_value(row, col)).strip() for col in range(sheet.ncols)
                            if str(sheet.cell_value(row, col)).strip()]
                    if vals:
                        print(" | ".join(vals))
            return
        except Exception as e:
            continue
    print("所有編碼均失敗")

for label, (path, ftype) in files.items():
    print(f"\n{'='*60}")
    print(f"【{label}】{os.path.basename(path)}")
    print('='*60)
    if not os.path.exists(path):
        print("找不到檔案")
        continue
    if ftype == "xlsx":
        read_xlsx(path)
    else:
        read_xls(path)
