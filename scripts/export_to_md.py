import io
import sys
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import openpyxl

TOKEN_PATH = r'C:\Users\kevin\.credentials\token.json'
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
FILE_ID = '1nvY4Axa1xTC_I3qcYa4qt6GS4ghOuUsu'
OUTPUT_PATH = r'C:\Users\kevin\Desktop\Taospace\output\立德班晚會分工_115年.md'

creds = Credentials.from_authorized_user_file(TOKEN_PATH, SCOPES)
drive = build('drive', 'v3', credentials=creds)

request = drive.files().get_media(fileId=FILE_ID)
buf = io.BytesIO()
downloader = MediaIoBaseDownload(buf, request)
done = False
while not done:
    status, done = downloader.next_chunk()
    print(f"下載進度：{int(status.progress() * 100)}%")

buf.seek(0)
wb = openpyxl.load_workbook(buf, data_only=True)

def sheet_to_md(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            rows.append(row)
    if not rows:
        return "（此工作表無資料）\n"

    # 找最大欄數
    max_cols = max(len(r) for r in rows)

    def fmt(val):
        if val is None:
            return ""
        return str(val).replace("\n", " ").replace("|", "｜")

    lines = []
    header = rows[0]
    lines.append("| " + " | ".join(fmt(c) for c in header) + " |")
    lines.append("| " + " | ".join("---" for _ in header) + " |")
    for row in rows[1:]:
        padded = list(row) + [None] * (max_cols - len(row))
        lines.append("| " + " | ".join(fmt(c) for c in padded) + " |")
    return "\n".join(lines) + "\n"

md_lines = [
    "# 2026立德班晚會分工",
    "",
    "> 來源：Google Drive `2026立德班晚會分工.xlsx`  ",
    "> 整理日期：2026-06-12",
    "",
]

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    md_lines.append(f"## {sheet_name}")
    md_lines.append("")
    md_lines.append(sheet_to_md(ws))

output = "\n".join(md_lines)

import os
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    f.write(output)

print(f"\n已儲存：{OUTPUT_PATH}")
