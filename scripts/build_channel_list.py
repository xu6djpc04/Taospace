# -*- coding: utf-8 -*-
"""產出 115年立德班晚會 音控輸入表 (channel list) xlsx
對齊實際 WING scene：
  Local 1-8 (CH1-8) / SD16 AES50 A (CH9-20) / 第二台箱 AES50 B (CH21-30,軌位待重排)
  / USB 播放 (CH31-32) / FX·RTN (CH37-40)。只錄影不直播(ZOOM不啟用)。"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = os.path.join(os.path.dirname(__file__), "115_立德班晚會_音控輸入表.xlsx")

wb = Workbook()

HDR_FILL = PatternFill("solid", fgColor="305496")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
NOTE_FONT = Font(italic=True, size=10, color="666666")
BLOCK_FONT = Font(bold=True, size=10, color="444444")
TODO_FONT = Font(bold=True, size=10, color="C00000")
thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

GROUP_FILL = {
    "戲劇": PatternFill("solid", fgColor="FCE4D6"),
    "樂團": PatternFill("solid", fgColor="DDEBF7"),
    "共用": PatternFill("solid", fgColor="E2EFDA"),
    "備用": PatternFill("solid", fgColor="F2F2F2"),
    "待定": PatternFill("solid", fgColor="FFF2CC"),
}
BLOCK_FILL = PatternFill("solid", fgColor="D9E1F2")

def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER

# ========== Sheet1：輸入表 ==========
ws = wb.active
ws.title = "輸入表(Channel List)"

ws["A1"] = "115年立德班晚會 — 音控輸入表（對齊 WING scene）"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:G1")
ws["A2"] = "Local 1-8 / SD16(AES50 A) CH9-20〔本場不接〕/ 第二台SD16(AES50 B) CH21-30=B1-B10：6耳掛+4樂器〔軌位待重排〕/ USB CH31-32 / FX·RTN CH37-40　｜　只錄影不直播"
ws["A2"].font = NOTE_FONT
ws.merge_cells("A2:G2")

headers = ["WING 軌", "scene 名稱", "patch 來源", "訊號源／角色", "節目組", "麥種／連接", "備註"]
hrow = 4
for i, h in enumerate(headers, 1):
    ws.cell(row=hrow, column=i, value=h)
style_header(ws, hrow, len(headers))

data = [
    ("block", "本機 Local（WING 8 in）"),
    ("row", "CH1", "MIC 1", "LCL 1", "主唱 A", "樂團", "無線手持麥", "✔ 已定"),
    ("row", "CH2", "MIC 2", "LCL 2", "主唱 B", "樂團", "無線手持麥", "✔ 已定"),
    ("row", "CH3", "MIC 3", "LCL 3", "記者", "戲劇", "無線手持麥", "✔ 已定"),
    ("row", "CH4", "MIC 4", "LCL 4", "旁白", "戲劇", "無線手持麥", "✔ 已定"),
    ("row", "CH5", "STAGE-01", "LCL 5", "備用／機動", "備用", "有線麥", ""),
    ("row", "CH6", "STAGE-02", "LCL 6", "備用／機動", "備用", "有線麥", ""),
    ("row", "CH7", "STAGE-03", "LCL 7", "備用／機動", "備用", "有線麥", ""),
    ("row", "CH8", "STAGE-04", "LCL 8", "備用／機動", "備用", "有線麥", ""),
    ("block", "SD16（AES50 A）CH9-20 — 本場不接"),
    ("row", "CH9-20", "SD16-01~12", "A 1-A 12", "（本場不接）", "備用", "—", "本場不接"),
    ("block", "第二台 SD16（AES50 B）CH21-30 = B1-B10：6 耳掛 ＋ 4 樂器〔軌位待重排，B1-B10 為前次設定〕"),
    ("row", "(B?待定)", "—", "AES50 B", "學長", "戲劇", "無線耳掛麥", "B1-B10 內，軌位待重排"),
    ("row", "(B?待定)", "—", "AES50 B", "學姐", "戲劇", "無線耳掛麥", "B1-B10 內，軌位待重排"),
    ("row", "(B?待定)", "—", "AES50 B", "學弟", "戲劇", "無線耳掛麥", "B1-B10 內，軌位待重排"),
    ("row", "(B?待定)", "—", "AES50 B", "機器人", "戲劇", "無線耳掛麥", "變聲／效果待確認"),
    ("row", "(B?待定)", "—", "AES50 B", "耳麥 5（角色待確認）", "待定", "無線耳掛麥", "6 支耳掛都接"),
    ("row", "(B?待定)", "—", "AES50 B", "耳麥 6（角色待確認）", "待定", "無線耳掛麥", "6 支耳掛都接"),
    ("row", "(B?待定)", "—", "AES50 B", "小提琴", "樂團", "收音麥／DI", "軌位待重排"),
    ("row", "(B?待定)", "—", "AES50 B", "木箱鼓", "樂團", "收音麥(動圈)", "軌位待重排"),
    ("row", "(B?待定)", "—", "AES50 B", "電子琴", "樂團", "DI", "走 DI；軌位待重排"),
    ("row", "(B?待定)", "—", "AES50 B", "電吉他", "樂團", "DI", "走 DI；軌位待重排"),
    ("block", "USB 播放（筆電）"),
    ("row", "CH31", "USB 1/2", "USB 1/2", "播放筆電（伴奏・音效）", "共用", "USB", ""),
    ("row", "CH32", "USB 9/10", "USB 9/10", "備用／其他播放", "共用", "USB", ""),
    ("block", "FX／群組回送"),
    ("row", "CH37", "Reverb", "FX1 (HALL)", "人聲殘響回送", "—", "FX", ""),
    ("row", "CH38", "Delay", "FX2 (ST-DL)", "人聲延遲回送", "—", "FX", ""),
    ("row", "CH39", "MIC RTN", "BUS 1", "麥克風群組回送", "—", "—", ""),
    ("row", "CH40", "PGM RTN", "BUS 2", "節目群組回送", "—", "—", ""),
]

r = hrow + 1
for item in data:
    if item[0] == "block":
        ws.cell(row=r, column=1, value=item[1])
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
        c = ws.cell(row=r, column=1)
        c.font = TODO_FONT if "待重排" in item[1] else BLOCK_FONT
        c.fill = BLOCK_FILL
        c.alignment = LEFT
        for cc in range(1, len(headers) + 1):
            ws.cell(row=r, column=cc).border = BORDER
    else:
        vals = item[1:]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            cell.alignment = LEFT if c in (4, 6, 7) else CENTER
        fill = GROUP_FILL.get(vals[4])
        if fill:
            ws.cell(row=r, column=5).fill = fill
    r += 1

widths = [10, 12, 11, 20, 7, 14, 20]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w
ws.freeze_panes = "A5"

# ========== Sheet2：AUX返送 / 匯流排 / 輸出 ==========
ws2 = wb.create_sheet("AUX・匯流排・輸出")
ws2["A1"] = "AUX 返送 / 匯流排 / 實體輸出"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A1:D1")

# AUX
ws2["A3"] = "AUX 返送"
ws2["A3"].font = BLOCK_FONT
ws2["A3"].fill = BLOCK_FILL
ws2.merge_cells("A3:D3")
aux_head = ["AUX", "scene 名稱", "patch 來源", "備註"]
for i, h in enumerate(aux_head, 1):
    ws2.cell(row=4, column=i, value=h)
style_header(ws2, 4, len(aux_head))
aux_rows = [
    ("A1", "ZOOM", "AUX 1/2", "★ 本場不啟用（只錄影、不直播）"),
    ("A2", "MUSIC-01", "AUX 3/4", "電腦 A（音樂）"),
    ("A3", "MUSIC-02", "AUX 5/6", "電腦 B（音樂）；不走 HDMI"),
    ("A4", "（空）", "—", "不走 HDMI，未使用"),
    ("A5", "SD16-13.14", "A 13/14", "SD16 立體聲返送"),
    ("A6", "SD16-15.16", "A 15/16", "SD16 立體聲返送"),
    ("A7 / A8", "（空）", "—", "未使用"),
]
r = 5
for row in aux_rows:
    for c, val in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.border = BORDER
        cell.alignment = LEFT if c in (2, 4) else CENTER
    r += 1

def render_table(r, title, head, rows, left_cols):
    ws2.cell(row=r, column=1, value=title).font = BLOCK_FONT
    ws2.cell(row=r, column=1).fill = BLOCK_FILL
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    for i, h in enumerate(head, 1):
        ws2.cell(row=r, column=i, value=h)
    style_header(ws2, r, len(head))
    r += 1
    for row in rows:
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = BORDER
            cell.alignment = LEFT if c in left_cols else CENTER
        r += 1
    return r + 1

# 群組匯流排 BUS
r = render_table(r + 1, "群組匯流排（BUS）", ["BUS", "名稱", "用途", "備註"], [
    ("B1", "MIC GUP", "麥克風群組", "彙整所有人聲（供錄音取音）"),
    ("B2", "PGM GUP", "節目群組", "節目總和（供錄音取音）"),
    ("B3", "LIVE MAIN", "現場主混", "現場主輸出"),
], (2, 3, 4))

# 主混 M1-M4
r = render_table(r, "主混匯流排（M1-M4）", ["匯流排", "名稱", "用途", "備註"], [
    ("M1", "MAIN SPK", "主擴聲", ""),
    ("M2", "ZONE 2", "觀眾席分區 2", ""),
    ("M3", "ZONE 3", "觀眾席分區 3", ""),
    ("M4", "ZONE 4", "觀眾席分區 4", ""),
], (2, 3, 4))

# 矩陣 MX1-MX8
r = render_table(r, "矩陣輸出（MX1-MX8）", ["矩陣", "名稱", "用途", "備註"], [
    ("MX1", "MAIN SPK", "主擴實體輸出", "主混經矩陣送主喇叭"),
    ("MX2", "STAGE M", "舞台監聽", "送 4 顆監聽（共用一路）"),
    ("MX3", "For ZOOM", "直播饋送", "★ 本場不啟用"),
    ("MX4-MX8", "（空）", "—", "未使用"),
], (2, 3, 4))

# 實體輸出對應
r = render_table(r, "實體輸出對應（功能）", ["輸出", "設備", "取音來源", "備註"], [
    ("主擴 L/R", "主動主喇叭 L/R", "MX1（MAIN SPK）", "主動式免外接功放"),
    ("重低音", "主動 Subwoofer", "主擴分頻", "由喇叭或 WING 分頻"),
    ("舞台監聽 ×4", "主動監聽 ×4", "MX2（STAGE M）", "四顆同一監聽混音"),
    ("觀眾席分區 ZONE 2-4", "觀眾席補聲喇叭", "M2／M3／M4", "實體接點依現場確認"),
    ("錄音", "WING REC（內錄 USB／SD）", "PGM GUP／LIVE MAIN", "錄音用 REC；如另有錄影機再拉乾淨混音"),
], (2, 3, 4))

for i, w in enumerate([14, 16, 18, 34], 1):
    ws2.column_dimensions[chr(64 + i)].width = w

# ========== Sheet3：待確認 / 注意事項 ==========
ws3 = wb.create_sheet("待確認與注意事項")
notes = [
    "115年立德班晚會 音控 — 待確認與注意事項",
    "",
    "【已鎖定】",
    "• 本機 Local CH1-4：無線手持麥 ×4 = 主唱A／主唱B／記者／旁白。",
    "• 本機 Local CH5-8：有線麥 ×4（備用／機動）。",
    "• 舞台訊源全走第二台 SD16（AES50 B）= CH21-30 = B1-B10：6 支無線耳掛 ＋ 4 件樂器，共 10 軌。",
    "• SD16（AES50 A，CH9-20）本場不接。",
    "• 音控室 2 台電腦：電腦 A → MUSIC-01（AUX A2）；電腦 B → MUSIC-02（AUX A3）；不走 HDMI。",
    "• 主混 M2/M3/M4 = ZONE 2/3/4 = 觀眾席分區。",
    "• 錄音用 WING REC（內錄）。只錄影、不直播（ZOOM/AUX A1 不啟用）。",
    "",
    "【待確認】",
    "1. AES50 B（B1-B10）軌位待重排 —— 目前 B1-B10 是前次活動設定；本場 6 耳掛＋4 樂器要重新指定各接哪一軌。",
    "2. 無線耳掛 6 支都接：學長／學姐／學弟／機器人 已定，另 2 支角色待確認。",
    "3. 電子琴、電吉他走 DI（需 DI Box，建議備 2 顆）；小提琴／木箱鼓走收音麥。",
    "4. 機器人聲音是否需變聲／效果（該軌加 FX）。",
    "5. 主擴／監聽／Sub 的實體輸出接點，依現場 SD16／控台 Out 確認。",
    "",
    "【監聽 / 輸出 / 錄音】",
    "• 輸出結構：M1 MAIN SPK／M2-4 觀眾席分區ZONE 2-4；MX1 MAIN SPK（主擴）／MX2 STAGE M（監聽）／MX3 For ZOOM（不啟用）。",
    "• 舞台監聽 4 顆共用一路（MX2 STAGE M，同訊號）；如戲劇與樂團需求差異大可再分路。",
    "• 錄音用 WING REC，取 PGM GUP／LIVE MAIN 乾淨混音；避免直接接 Main 受場地回授影響。",
    "• 戲劇↔樂團換場：先靜音(mute)無線麥群組再換人，留意監聽音量。",
]
for i, line in enumerate(notes, 1):
    cell = ws3.cell(row=i, column=1, value=line)
    cell.alignment = LEFT
    if i == 1:
        cell.font = TITLE_FONT
    elif line.startswith("【"):
        cell.font = BLOCK_FONT
ws3.column_dimensions["A"].width = 100

wb.save(OUT)
print("OK ->", OUT)
